#!/usr/bin/env python3
"""
00-buildPhosphositeFile.py: build the phosphosites reference CSV from three sources:

  1. Ochoa et al. (2020) Nat Biotechnol 38, 365-373: ~116k curated human
     phosphosites from Supplementary Table S3 (Springer static content).
     https://static-content.springer.com/esm/art%3A10.1038%2Fs41587-019-0344-3/
         MediaObjects/41587_2019_344_MOESM4_ESM.xlsx

  2. UniProt PTM annotations: all reviewed human proteins with phospho
     modifications from the UniProt REST API (~41k sites, ~12k unique to UniProt).

  3. Synapse supplement(s): raw phosphoproteomics CSV/TSV files on Synapse
     whose site column is parsed to add experiment-specific sites not in the
     databases above (--synapse_supplements syn70078415 ...).

Sources 1 and 2 are always fetched. Source 3 is optional but recommended.

UniProt accession -> gene_symbol: UniProt REST API (TSV, accession+gene_names).
gene_symbol -> entrez_id: genes.csv.

Site notation: <GENE>-<Residue><Position><mod_code>
    e.g. AAAS-S495s  (s=phosphoserine, t=phosphothreonine, y=phosphotyrosine)

Output columns:
    phosphosite_id, entrez_id, gene_symbol, residue, position, modification, other_id

Usage:
    python 00-buildPhosphositeFile.py <genes.csv>
                                      [--out /tmp/phosphosites.csv]
                                      [--prev /tmp/prev_phosphosites.csv]
                                      [--supplement site_file.csv ...]
                                      [--site_col site]
                                      [--synapse_supplements syn70078415 ...]
                                      [--synapse_site_col site]
                                      [--ochoa_url <url>]
                                      [--uniprot_gene_url <url>]
                                      [--uniprot_ptm_url <url>]
"""

import argparse
import csv
import io
import json
import logging
import re
import sys
import tempfile
import time
import urllib.request

import openpyxl
import pandas as pd


OCHOA_URL = (
    "https://static-content.springer.com/esm/"
    "art%3A10.1038%2Fs41587-019-0344-3/MediaObjects/"
    "41587_2019_344_MOESM4_ESM.xlsx"
)

UNIPROT_GENE_URL = (
    "https://rest.uniprot.org/uniprotkb/search"
    "?query=reviewed:true+AND+organism_id:9606"
    "&fields=accession,gene_names"
    "&format=tsv"
    "&size=500"
)

UNIPROT_PTM_URL = (
    "https://rest.uniprot.org/uniprotkb/search"
    "?query=reviewed:true+AND+organism_id:9606+AND+ft_mod_res:phospho*"
    "&fields=accession,gene_names,ft_mod_res"
    "&format=json"
    "&size=200"
)

_RESIDUE_MOD = {"S": "s", "T": "t", "Y": "y"}
_PHOSPHO_DESC_MAP = {
    "phosphoserine":    ("S", "s"),
    "phosphothreonine": ("T", "t"),
    "phosphotyrosine":  ("Y", "y"),
}
_SITE_RE = re.compile(r'^(.+)-([STY])(\d+)([a-z]+)$', re.IGNORECASE)
_SUPP_MOD_CODES = {'s': 'phospho', 't': 'phospho', 'y': 'phospho'}

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0"
    ),
    "Accept": "*/*",
}


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
def configure_logging():
    logging.basicConfig(
        format="[%(asctime)s] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        level=logging.INFO,
    )


# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------
def load_gene_map(genes_path: str) -> pd.DataFrame:
    df = pd.read_csv(genes_path)
    needed = {'gene_symbol', 'entrez_id'}
    if needed - set(df.columns):
        raise KeyError(f"genes file missing columns: {needed - set(df.columns)}")
    df = df.dropna(subset=['gene_symbol', 'entrez_id']).copy()
    df['entrez_id'] = df['entrez_id'].astype(int)
    df = df.drop_duplicates(subset='gene_symbol')[['gene_symbol', 'entrez_id']]
    logging.info("Loaded %d gene_symbol → entrez_id mappings", len(df))
    return df


# ---------------------------------------------------------------------------
# Network helper
# ---------------------------------------------------------------------------
def _fetch_url(url: str, retries: int = 3, timeout: int = 120) -> bytes:
    req = urllib.request.Request(url, headers=_HEADERS)
    for attempt in range(1, retries + 1):
        try:
            logging.info("  GET %s (attempt %d/%d)", url, attempt, retries)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except Exception as exc:
            logging.warning("  Attempt %d failed: %s", attempt, exc)
            if attempt < retries:
                time.sleep(5 * attempt)
    raise RuntimeError(f"Failed to fetch {url} after {retries} attempts")


def _fetch_json_page(url: str, timeout: int = 120) -> tuple[list[dict], str | None]:
    req = urllib.request.Request(url, headers={**_HEADERS, "Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode('utf-8'))
        link_header = resp.headers.get('Link', '')
    entries = data.get('results', [])
    m = re.search(r'<([^>]+)>;\s*rel="next"', link_header)
    return entries, m.group(1) if m else None


# ---------------------------------------------------------------------------
# Source 1: Ochoa et al. supplementary table
# ---------------------------------------------------------------------------
def fetch_ochoa_phosphosites(url: str) -> list[dict]:
    """Download Ochoa MOESM4 xlsx and return raw phosphosite records."""
    logging.info("Downloading Ochoa et al. supplementary phosphoproteome")
    data = _fetch_url(url)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name

    logging.info("Parsing xlsx (%d MB) …", len(data) // 1_000_000)
    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

    sheet_name = "annotated_phosphoproteome"
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    records, header, uni_i, pos_i, res_i, skipped = [], None, None, None, None, 0

    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in row]
            try:
                uni_i = header.index("uniprot")
                pos_i = header.index("position")
                res_i = header.index("residue")
            except ValueError as exc:
                wb.close()
                raise ValueError(f"Required column missing in Ochoa sheet: {exc}")
            logging.info(
                "Ochoa columns: uniprot=%d, position=%d, residue=%d",
                uni_i, pos_i, res_i,
            )
            continue

        acc = str(row[uni_i]).strip() if row[uni_i] is not None else ""
        try:
            pos = int(row[pos_i])
        except (TypeError, ValueError):
            skipped += 1
            continue

        res_raw = str(row[res_i]).strip().upper() if row[res_i] is not None else ""
        residue = res_raw[0] if res_raw and res_raw[0] in _RESIDUE_MOD else None

        if residue is None or not acc:
            skipped += 1
            continue

        records.append({
            "accession":    acc,
            "residue":      residue,
            "position":     pos,
            "modification": "phospho",
            "mod_code":     _RESIDUE_MOD[residue],
        })

    wb.close()
    logging.info(
        "Ochoa: %d records parsed (%d skipped)", len(records), skipped
    )
    return records


# ---------------------------------------------------------------------------
# Source 2: UniProt PTM annotations
# ---------------------------------------------------------------------------
def fetch_uniprot_ptm_sites(url: str) -> list[dict]:
    """Fetch phosphorylation sites from UniProt Modified residue features.

    Returns records with gene_symbol (already resolved) instead of accession.
    """
    logging.info("Fetching UniProt PTM phosphorylation sites")
    all_records = []
    next_url: str | None = url
    page = 0

    while next_url:
        page += 1
        if page % 10 == 1:
            logging.info("  UniProt PTM page %d (%d records so far)", page, len(all_records))

        for attempt in range(1, 4):
            try:
                entries, next_url = _fetch_json_page(next_url)
                break
            except Exception as exc:
                logging.warning("  Page %d attempt %d failed: %s", page, attempt, exc)
                if attempt == 3:
                    raise
                time.sleep(5 * attempt)

        for entry in entries:
            gene = ""
            for g in entry.get("genes", []):
                if "geneName" in g:
                    gene = g["geneName"].get("value", "").upper()
                    break
            if not gene:
                continue
            for feat in entry.get("features", []):
                if feat.get("type") != "Modified residue":
                    continue
                desc = feat.get("description", "").lower()
                residue, mod_code = None, None
                for key, (aa, mc) in _PHOSPHO_DESC_MAP.items():
                    if desc.startswith(key):
                        residue, mod_code = aa, mc
                        break
                if residue is None:
                    continue
                pos = feat.get("location", {}).get("start", {}).get("value")
                if pos is None:
                    continue
                all_records.append({
                    "gene_symbol":  gene,
                    "residue":      residue,
                    "position":     int(pos),
                    "modification": "phospho",
                    "mod_code":     mod_code,
                    "other_id":     f"{gene}-{residue}{pos}{mod_code}",
                })

    logging.info(
        "UniProt PTM: %d records from %d pages", len(all_records), page
    )
    return all_records


# ---------------------------------------------------------------------------
# UniProt accession → gene_symbol
# ---------------------------------------------------------------------------
def fetch_uniprot_gene_map(url: str) -> dict[str, str]:
    """Return accession → primary gene_symbol for all reviewed human proteins."""
    logging.info("Fetching UniProt accession → gene_symbol map")
    acc_to_gene: dict[str, str] = {}
    next_url: str | None = url
    page = 0

    while next_url:
        page += 1
        if page % 10 == 1:
            logging.info(
                "  UniProt gene map page %d (%d entries so far)", page, len(acc_to_gene)
            )
        for attempt in range(1, 4):
            try:
                req = urllib.request.Request(next_url, headers=_HEADERS)
                with urllib.request.urlopen(req, timeout=120) as resp:
                    raw = resp.read()
                    link_header = resp.headers.get("Link", "")
                break
            except Exception as exc:
                logging.warning("  Page %d attempt %d failed: %s", page, attempt, exc)
                if attempt == 3:
                    raise
                time.sleep(5 * attempt)

        reader = csv.DictReader(
            io.StringIO(raw.decode("utf-8", errors="replace")), delimiter="\t"
        )
        for row in reader:
            acc = row.get("Entry", row.get("accession", "")).strip()
            gene_names = row.get("Gene Names", row.get("gene_names", "")).strip()
            if acc and gene_names:
                acc_to_gene[acc] = gene_names.split()[0].upper()

        m = re.search(r'<([^>]+)>;\s*rel="next"', link_header)
        next_url = m.group(1) if m else None

    logging.info(
        "UniProt gene map: %d entries (%d pages)", len(acc_to_gene), page
    )
    return acc_to_gene


# ---------------------------------------------------------------------------
# Source 3: Synapse supplement(s)
# ---------------------------------------------------------------------------
def fetch_synapse_supplement(syn_id: str, site_col: str) -> list[dict]:
    """Download a Synapse file and parse its site strings as supplement records."""
    import synapseclient
    logging.info("Fetching Synapse supplement %s", syn_id)
    syn = synapseclient.Synapse()
    syn.login(silent=True)
    entity = syn.get(syn_id)
    path = entity.path
    sep = "\t" if path.endswith((".tsv", ".tsv.gz")) else ","

    import gzip as _gzip
    open_fn = _gzip.open if path.endswith(".gz") else open

    records = []
    skipped = 0
    seen: set[str] = set()
    resolved_col: str | None = None

    with open_fn(path, "rt", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter=sep)
        for row in reader:
            # Auto-detect site column on first row if requested column missing
            if resolved_col is None:
                if site_col in row:
                    resolved_col = site_col
                else:
                    for candidate in ["site", "Site", "phosphosite", "feature_id"]:
                        if candidate in row:
                            resolved_col = candidate
                            logging.info(
                                "  Synapse %s: using column '%s' for sites",
                                syn_id, resolved_col,
                            )
                            break
                if resolved_col is None:
                    logging.warning(
                        "  Synapse %s: could not find site column (tried %s). "
                        "Available: %s",
                        syn_id, site_col, list(row.keys())[:10],
                    )
                    return []

            site = row.get(resolved_col, "").strip()
            if not site or site in seen:
                continue
            seen.add(site)
            m = _SITE_RE.match(site)
            if not m:
                skipped += 1
                continue
            mod_code = m.group(4).lower()
            records.append({
                "gene_symbol":  m.group(1).upper(),
                "residue":      m.group(2).upper(),
                "position":     int(m.group(3)),
                "modification": _SUPP_MOD_CODES.get(mod_code, f"phospho_{mod_code}"),
                "mod_code":     mod_code,
                "other_id":     site,
            })

    logging.info(
        "Synapse %s: %d unique sites parsed (%d skipped)", syn_id, len(records), skipped
    )
    return records


# ---------------------------------------------------------------------------
# Local file supplement
# ---------------------------------------------------------------------------
def parse_supplement_sites(path: str, site_col: str) -> list[dict]:
    records = []
    skipped = 0
    sep = "\t" if path.endswith(".tsv") else ","
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=sep)
        seen: set[str] = set()
        for row in reader:
            site = row.get(site_col, "").strip()
            if not site or site in seen:
                continue
            seen.add(site)
            m = _SITE_RE.match(site)
            if not m:
                skipped += 1
                continue
            mod_code = m.group(4).lower()
            records.append({
                "gene_symbol":  m.group(1).upper(),
                "residue":      m.group(2).upper(),
                "position":     int(m.group(3)),
                "modification": _SUPP_MOD_CODES.get(mod_code, f"phospho_{mod_code}"),
                "mod_code":     mod_code,
                "other_id":     site,
            })
    logging.info(
        "Supplement %s: %d unique sites (%d skipped)", path, len(records), skipped
    )
    return records


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------
def build_phosphosite_table(
    ochoa_records:   list[dict],
    acc_to_gene:     dict[str, str],
    uniprot_records: list[dict],
    gene_map:        pd.DataFrame,
    supplement_records: list[dict],
    prev: pd.DataFrame | None,
) -> pd.DataFrame:
    gene_to_entrez = gene_map.set_index("gene_symbol")["entrez_id"].to_dict()
    rows: list[dict] = []
    existing_ids: set[str] = set()

    def _add(gene: str, residue: str, position: int, modification: str, mod_code: str) -> None:
        other_id = f"{gene}-{residue}{position}{mod_code}"
        if other_id in existing_ids:
            return
        entrez = gene_to_entrez.get(gene)
        if entrez is None:
            return
        rows.append({
            "gene_symbol":  gene,
            "entrez_id":    entrez,
            "residue":      residue,
            "position":     position,
            "modification": modification,
            "other_id":     other_id,
        })
        existing_ids.add(other_id)

    # --- Ochoa (primary) ---
    no_gene_ochoa = 0
    for rec in ochoa_records:
        gene = acc_to_gene.get(rec["accession"], "")
        if not gene:
            no_gene_ochoa += 1
            continue
        _add(gene, rec["residue"], rec["position"], rec["modification"], rec["mod_code"])
    logging.info(
        "Ochoa: %d sites added (%d with no gene symbol)", len(rows), no_gene_ochoa
    )

    before_uniprot = len(rows)
    # --- UniProt PTM (secondary) ---
    for rec in uniprot_records:
        _add(
            rec["gene_symbol"], rec["residue"], rec["position"],
            rec["modification"], rec["mod_code"],
        )
    logging.info(
        "UniProt PTM: +%d new sites (total %d)", len(rows) - before_uniprot, len(rows)
    )

    before_supp = len(rows)
    # --- Supplements (local files + Synapse) ---
    for rec in supplement_records:
        _add(
            rec["gene_symbol"], rec["residue"], rec["position"],
            rec["modification"], rec["mod_code"],
        )
    if len(rows) > before_supp:
        logging.info(
            "Supplements: +%d new sites (total %d)", len(rows) - before_supp, len(rows)
        )

    df = pd.DataFrame(rows)

    # --- Assign / preserve phosphosite_ids ---
    if prev is not None and not prev.empty:
        prev_map = prev.set_index("other_id")["phosphosite_id"].to_dict()
        next_id = int(prev["phosphosite_id"].max()) + 1
    else:
        prev_map = {}
        next_id = 1

    ids = []
    for site in df["other_id"]:
        if site in prev_map:
            ids.append(prev_map[site])
        else:
            ids.append(next_id)
            prev_map[site] = next_id
            next_id += 1
    df["phosphosite_id"] = ids

    cols = [
        "phosphosite_id", "entrez_id", "gene_symbol",
        "residue", "position", "modification", "other_id",
    ]
    df = df[cols].sort_values("phosphosite_id").reset_index(drop=True)
    logging.info(
        "Final phosphosites table: %d rows, %d unique genes",
        len(df), df["gene_symbol"].nunique(),
    )
    return df


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("genes", help="genes.csv (gene_symbol → entrez_id)")
    parser.add_argument("--out",  default="/tmp/phosphosites.csv")
    parser.add_argument("--prev", default=None,
                        help="Previous phosphosites.csv to preserve IDs")
    parser.add_argument("--supplement", nargs="*", default=[],
                        help="Local CSV/TSV file(s) with site strings")
    parser.add_argument("--site_col", default="site",
                        help="Column name for site strings in local supplement files")
    parser.add_argument("--synapse_supplements", nargs="*", default=[],
                        help="Synapse IDs of raw phospho files to add as supplements "
                             "(e.g. syn70078415)")
    parser.add_argument("--synapse_site_col", default="site",
                        help="Site column name in Synapse supplement files")
    parser.add_argument("--ochoa_url",       default=OCHOA_URL)
    parser.add_argument("--uniprot_gene_url", default=UNIPROT_GENE_URL)
    parser.add_argument("--uniprot_ptm_url",  default=UNIPROT_PTM_URL)
    args = parser.parse_args()

    configure_logging()

    gene_map = load_gene_map(args.genes)

    prev = None
    if args.prev:
        prev = pd.read_csv(args.prev)
        logging.info("Loaded %d previous phosphosites from %s", len(prev), args.prev)

    ochoa_records   = fetch_ochoa_phosphosites(args.ochoa_url)
    acc_to_gene     = fetch_uniprot_gene_map(args.uniprot_gene_url)
    uniprot_records = fetch_uniprot_ptm_sites(args.uniprot_ptm_url)

    supp_records: list[dict] = []
    for path in (args.supplement or []):
        supp_records.extend(parse_supplement_sites(path, args.site_col))
    for syn_id in (args.synapse_supplements or []):
        try:
            supp_records.extend(fetch_synapse_supplement(syn_id, args.synapse_site_col))
        except Exception as exc:
            logging.warning("Synapse supplement %s failed: %s — skipping", syn_id, exc)

    df = build_phosphosite_table(
        ochoa_records, acc_to_gene, uniprot_records, gene_map, supp_records, prev
    )

    df.to_csv(args.out, index=False)
    logging.info(
        "Wrote %s (%d rows, %d unique genes)",
        args.out, len(df), df["gene_symbol"].nunique(),
    )


if __name__ == "__main__":
    sys.exit(main() or 0)
